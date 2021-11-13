import json
import openpyxl
from pathlib import Path
from account_promotion.utils import *

with open("output/accounts.txt","r") as f:
    users = json.load(f)
oname = "graph.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Vertices"
ws["A1"] = "Id"
usr = 2
final_ids = []
for user in users:
    id1 = user["id"]
    final_ids += [id1]
    ws["A"+str(usr)] = str(id1)
    usr += 1

ws = wb.create_sheet("Edges")
ws["A1"] = "Source"
ws["B1"] = "Target"
ws["C1"] = "Type"
edge = 2
for user in users:
    id1 = user["id"]
    for id2 in user["friends"]:
        if (id2 in final_ids):
            ws["A"+str(edge)] = str(id1)
            ws["B"+str(edge)] = str(id2)
            ws["C"+str(edge)] = "Undirected"
            edge += 1

wb.save(oname)
